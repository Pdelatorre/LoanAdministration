import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List
import config
import os


def generate_audit_report(
    loan,
    schedule: List[Dict],
    loan_id: str,
    output_dir: str = None
) -> str:
    """
    Generate comprehensive Excel audit report.
    
    Creates multi-tab workbook with:
    - Loan summary
    - Period detail
    - Investor allocations
    - Payment ledger
    - Ownership history
    - Reconciliation
    
    Args:
        loan: Loan object
        schedule: Complete loan schedule
        loan_id: Loan identifier
        output_dir: Where to save report
    
    Returns:
        Path to generated Excel file
    """
    if output_dir is None:
        output_dir = config.AUDIT_REPORTS_DIR
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Tab 1: Loan Summary
    _create_loan_summary_tab(wb, loan, schedule)

    # Tab 2: Interest Period Detail
    _create_period_detail_tab(wb, schedule)

    # Tab 3: Investor Allocations
    _create_investor_allocations_tab(wb, loan_id, schedule)

    # Tab 4: Investor Roll-Forward
    _create_investor_roll_forward_tab(wb, loan_id, schedule)

    # Tab 5: Payment Ledger
    _create_payment_ledger_tab(wb, loan_id)

    # Tab 6: Fee Income
    _create_fee_income_tab(wb, loan_id)

    # Tab 7: Ownership History
    _create_ownership_history_tab(wb, loan_id)

    # Tab 8: Reconciliation Checks
    _create_reconciliation_tab(wb, loan_id, schedule)
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d')
    filename = f"{loan.loan_name}_Audit_Report_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    print(f"✅ Generated audit report: {filename}")
    return filepath


def _create_loan_summary_tab(wb, loan, schedule):
    """Create Loan Summary tab."""
    ws = wb.create_sheet("Loan Summary")
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = "LOAN SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Loan details
    row = 3
    convention_label = "Last Business Day" if loan.period_end_convention == "last_business_day" else "Calendar Month End"

    details = [
        ("Loan ID:", loan.loan_id),
        ("Borrower:", loan.borrower),
        ("Loan Name:", loan.loan_name),
        ("Principal Amount:", f"${loan.principal:,.2f}"),
        ("Margin:", f"{loan.margin * 100:.2f}%"),
        ("SOFR Floor:", f"{loan.sofr_floor * 100:.2f}%"),
        ("Origination Date:", loan.origination_date.strftime('%Y-%m-%d')),
        ("Maturity Date:", loan.maturity_date.strftime('%Y-%m-%d')),
        ("Period End Convention:", convention_label),
        ("Day Count Convention:", "Actual/360"),
        ("", ""),
        ("Total Periods:", len(schedule)),
        ("Total Interest:", f"${sum(p['interest_owed'] for p in schedule):,.2f}"),
        ("Interest Prepaid at Close:", f"${loan.interest_prepayment:,.2f}"),
        ("Prepaid Interest Applied:", f"${sum(p.get('prepaid_applied', 0) for p in schedule):,.2f}"),
        ("Current Principal:", f"${schedule[-1]['principal_ending']:,.2f}"),
    ]

    # OID funding waterfall — only shown when OID is present
    oid_amount = getattr(loan, 'oid_amount', 0)
    if oid_amount > 0:
        from oid_calculations import compute_net_investor_call, compute_net_borrower_advance
        closing_expenses = getattr(loan, 'closing_expenses', 0)
        net_call = compute_net_investor_call(loan.principal, loan.interest_prepayment, oid_amount)
        net_adv  = compute_net_borrower_advance(net_call, closing_expenses)
        total_oid_recognized = sum(p.get('period_oid', 0) for p in schedule)
        details += [
            ("", ""),
            ("── OID FUNDING WATERFALL ──", ""),
            ("OID Amount:", f"${oid_amount:,.2f}"),
            ("Interest Prepaid at Close:", f"${loan.interest_prepayment:,.2f}"),
            ("Net Investor Call:", f"${net_call:,.2f}"),
        ]
        if closing_expenses > 0:
            details.append(("Closing Expenses:", f"${closing_expenses:,.2f}"))
        details += [
            ("Net Borrower Advance:", f"${net_adv:,.2f}"),
            ("", ""),
            ("Total OID Recognized to Date:", f"${total_oid_recognized:,.2f}"),
            ("OID Unamortized (Contra-Asset):", f"${schedule[-1].get('oid_unamortized_end', 0):,.2f}"),
        ]
    
    for label, value in details:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def _create_period_detail_tab(wb, schedule):
    """Create Interest Period Detail tab."""
    ws = wb.create_sheet("Period Detail")

    # Detect OID — add columns only when the loan has OID
    has_oid = any(p.get('period_oid', 0) != 0 for p in schedule)

    # Headers
    headers = [
        "Period", "Start Date", "End Date", "Days",
        "Principal Beginning", "SOFR Rate", "Margin", "Effective Rate",
        "Interest Owed",
    ]
    if has_oid:
        headers += ["Period OID", "OID Unamortized End"]
    headers += [
        "Interest Paid", "Prepaid Applied", "Prepaid Balance End",
        "Principal Ending", "Status"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, period in enumerate(schedule, 2):
        ws.cell(row=row, column=1, value=period['period_number'])
        ws.cell(row=row, column=2, value=period['start_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=period['end_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value=period['days'])
        ws.cell(row=row, column=5, value=period['principal_beginning'])
        ws.cell(row=row, column=6, value=period.get('sofr_rate', 0))
        ws.cell(row=row, column=7, value=period.get('margin', 0))
        ws.cell(row=row, column=8, value=period.get('effective_rate', 0))
        ws.cell(row=row, column=9, value=period['interest_owed'])

        # OID columns (shift subsequent columns when present)
        col_offset = 0
        if has_oid:
            ws.cell(row=row, column=10, value=period.get('period_oid', 0))
            ws.cell(row=row, column=11, value=period.get('oid_unamortized_end', 0))
            col_offset = 2

        ws.cell(row=row, column=10 + col_offset, value=period.get('amount_paid', 0))
        ws.cell(row=row, column=11 + col_offset, value=period.get('prepaid_applied', 0))
        ws.cell(row=row, column=12 + col_offset, value=period.get('prepaid_balance_end', 0))
        ws.cell(row=row, column=13 + col_offset, value=period['principal_ending'])

        # Determine status based on payment
        payment_status = period.get('payment_status')
        if payment_status == 'Paid':
            status = 'Final'
        elif payment_status == 'Partially Paid':
            status = 'Partially Paid'
        elif payment_status == 'Unpaid':
            status = 'Unpaid'
        else:
            status = 'Projected'

        ws.cell(row=row, column=14 + col_offset, value=status)

        # Format currency columns
        currency_cols = [5, 9, 10 + col_offset, 11 + col_offset, 12 + col_offset, 13 + col_offset]
        if has_oid:
            currency_cols += [10, 11]  # period_oid, oid_unamortized_end
        for col in currency_cols:
            ws.cell(row=row, column=col).number_format = '$#,##0.00'

        # Format percentage columns
        for col in [6, 7, 8]:
            ws.cell(row=row, column=col).number_format = '0.00000%'

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_investor_allocations_tab(wb, loan_id, schedule):
    """Create Investor Allocations tab."""
    from investors import load_investors
    from investor_allocation import allocate_period_to_investors
    
    ws = wb.create_sheet("Investor Allocations")
    
    # Headers
    headers = [
        "Period", "Investor ID", "Investor Name", "Ownership %",
        "Principal Begin", "Interest Income", "Prepayment", "Principal End"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row = 2
    for period in schedule:
        try:
            allocation = allocate_period_to_investors(loan_id, period)
            
            for investor in allocation['investor_allocations']:
                ws.cell(row=row, column=1, value=period['period_number'])
                ws.cell(row=row, column=2, value=investor['investor_id'])
                ws.cell(row=row, column=3, value=investor['investor_name'])
                
                # Get ownership % from last segment
                last_segment = allocation['ownership_segments'][-1]
                ownership = next(
                    (inv['ownership_pct'] for inv in last_segment['investors'] 
                     if inv['investor_id'] == investor['investor_id']),
                    0.0
                )
                ws.cell(row=row, column=4, value=ownership / 100)
                
                ws.cell(row=row, column=5, value=investor['principal_beginning'])
                ws.cell(row=row, column=6, value=investor['interest'])
                ws.cell(row=row, column=7, value=investor['principal_prepayment'])
                ws.cell(row=row, column=8, value=investor['principal_ending'])
                
                # Format currency columns
                for col in [5, 6, 7, 8]:
                    ws.cell(row=row, column=col).number_format = '$#,##0.00'
                
                # Format percentage
                ws.cell(row=row, column=4).number_format = '0.00%'
                
                row += 1
        except:
            # No investors for this loan
            pass
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _create_investor_roll_forward_tab(wb, loan_id, schedule):
    """
    Create Investor Roll-Forward tab.

    Accounting-style roll-forward for each investor showing:
      - Opening principal balance (loan origination)
      - One row per interest period: period dates, ownership %, cash interest,
        PIK interest, OID accreted, fees, principal prepayments, ending principal
      - Running cumulative columns for income, OID, and fees
      - Subtotal row per investor (bold, shaded)
      - Grand-total cross-check block at the bottom

    Layout:
      Col A  Period #
      Col B  Period Start
      Col C  Period End
      Col D  Ownership % (period-end)
      Col E  Beginning Principal
      Col F  Cash Interest
      Col G  PIK Interest
      Col H  Total Interest Income
      Col I  OID Accreted          (hidden/zero when no OID on loan)
      Col J  Fee Income
      Col K  Total Income
      Col L  Principal Prepayment
      Col M  Ending Principal
      Col N  Cumulative Interest
      Col O  Cumulative OID         (hidden/zero when no OID)
      Col P  Cumulative Fees
      Col Q  Cumulative Total Income
    """
    from investor_allocation import allocate_period_to_investors
    from fee_allocation import calculate_investor_fee_totals

    ws = wb.create_sheet("Investor Roll-Forward")

    # ── Styling helpers ──────────────────────────────────────────────────────
    BLUE       = "366092"
    LIGHT_BLUE = "D9E1F2"
    SUBTOTAL   = "BDD7EE"
    GRAND_FILL = "1F3864"

    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill(start_color=BLUE,    end_color=BLUE,    fill_type="solid")
    sub_font   = Font(bold=True, size=10)
    sub_fill   = PatternFill(start_color=SUBTOTAL, end_color=SUBTOTAL, fill_type="solid")
    alt_fill   = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    grand_font = Font(bold=True, color="FFFFFF", size=10)
    grand_fill = PatternFill(start_color=GRAND_FILL, end_color=GRAND_FILL, fill_type="solid")

    CURRENCY = '$#,##0.00'
    PERCENT  = '0.00%'

    # ── Detect OID ───────────────────────────────────────────────────────────
    has_oid = any(p.get('period_oid', 0) != 0 for p in schedule)

    # ── Column definitions (label, width) ────────────────────────────────────
    columns = [
        ("Period",             9),
        ("Period Start",      13),
        ("Period End",        13),
        ("Ownership %",       12),
        ("Beg. Principal",    16),
        ("Cash Interest",     15),
        ("PIK Interest",      14),
        ("Total Interest",    15),
        ("OID Accreted",      14),
        ("Fee Income",        13),
        ("Total Income",      14),
        ("Principal Prepay",  16),
        ("End. Principal",    16),
        ("Cumul. Interest",   16),
        ("Cumul. OID",        13),
        ("Cumul. Fees",       13),
        ("Cumul. Total Inc.", 17),
    ]

    # Column indices (1-based)
    COL = {label: idx for idx, (label, _) in enumerate(columns, 1)}

    def _set_col_widths():
        for idx, (_, width) in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _write_header_row(row):
        for idx, (label, _) in enumerate(columns, 1):
            c = ws.cell(row=row, column=idx, value=label)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def _fmt_currency(row, col):
        ws.cell(row=row, column=col).number_format = CURRENCY

    def _fmt_pct(row, col):
        ws.cell(row=row, column=col).number_format = PERCENT

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="INVESTOR ROLL-FORWARD — INCOME & BALANCE SUMMARY").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.row_dimensions[1].height = 20

    # ── Build per-period allocation cache (avoid recomputing per investor) ───
    period_allocations = []
    for period in schedule:
        try:
            alloc = allocate_period_to_investors(loan_id, period)
        except Exception:
            alloc = None
        period_allocations.append((period, alloc))

    # ── Collect all unique investors (preserve order of first appearance) ────
    seen = {}
    for period, alloc in period_allocations:
        if alloc is None:
            continue
        for inv in alloc['investor_allocations']:
            iid = inv['investor_id']
            if iid not in seen:
                seen[iid] = inv['investor_name']
    all_investor_ids = list(seen.keys())

    if not all_investor_ids:
        ws.cell(row=3, column=1, value="No investor data available for this loan.")
        _set_col_widths()
        return

    # ── Grand-total accumulators ──────────────────────────────────────────────
    grand = {
        'cash_interest': 0.0, 'pik_interest': 0.0, 'total_interest': 0.0,
        'oid': 0.0, 'fees': 0.0, 'total_income': 0.0, 'prepayments': 0.0,
    }

    current_row = 3   # leave row 2 blank as spacer

    # ── One block per investor ────────────────────────────────────────────────
    for inv_idx, investor_id in enumerate(all_investor_ids):
        investor_name = seen[investor_id]

        # Investor heading
        ws.cell(row=current_row, column=1, value=f"  {investor_name}  ({investor_id})").font = Font(bold=True, size=11)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(columns))
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Column header row
        _write_header_row(current_row)
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # ── Per-period data rows ──────────────────────────────────────────────
        inv_totals = {
            'cash_interest': 0.0, 'pik_interest': 0.0, 'total_interest': 0.0,
            'oid': 0.0, 'fees': 0.0, 'total_income': 0.0, 'prepayments': 0.0,
        }
        cumul_interest = 0.0
        cumul_oid      = 0.0
        cumul_fees     = 0.0
        cumul_total    = 0.0
        first_beg_principal = None
        last_end_principal  = None

        for p_idx, (period, alloc) in enumerate(period_allocations):
            if alloc is None:
                continue

            # Find this investor in the allocation
            inv_data = next(
                (i for i in alloc['investor_allocations'] if i['investor_id'] == investor_id),
                None
            )

            # OID for this investor this period
            period_oid_total = period.get('period_oid', 0.0)
            oid_unamortized_start = period.get('oid_unamortized_start',
                                               period.get('oid_unamortized_end', 0.0) + period_oid_total)
            if has_oid and inv_data and period_oid_total > 0:
                # Determine ownership ratio using last segment
                last_seg = alloc['ownership_segments'][-1]
                ownership_ratio = next(
                    (i['ownership_pct'] / 100.0 for i in last_seg['investors']
                     if i['investor_id'] == investor_id), 0.0
                )
                investor_oid = round(period_oid_total * ownership_ratio, 2)
            else:
                investor_oid = 0.0

            # Fees for this investor this period
            try:
                fee_result = calculate_investor_fee_totals(loan_id, period['period_number'], investor_id)
                investor_fees = fee_result['total_fees']
            except Exception:
                investor_fees = 0.0

            # Values (zeros if investor not present this period)
            if inv_data:
                cash_int  = inv_data.get('cash_interest', inv_data['interest'])
                pik_int   = inv_data.get('pik_interest', 0.0)
                tot_int   = inv_data['interest']
                beg_prin  = inv_data['principal_beginning']
                end_prin  = inv_data['principal_ending']
                prepay    = inv_data['principal_prepayment']

                # Get period-end ownership %
                last_seg  = alloc['ownership_segments'][-1]
                ownership = next(
                    (i['ownership_pct'] for i in last_seg['investors']
                     if i['investor_id'] == investor_id), 0.0
                )
            else:
                cash_int = pik_int = tot_int = beg_prin = end_prin = prepay = ownership = 0.0

            total_income = tot_int + investor_oid + investor_fees

            # Running cumulative
            cumul_interest += tot_int
            cumul_oid      += investor_oid
            cumul_fees     += investor_fees
            cumul_total    += total_income

            # Track investor-level totals
            inv_totals['cash_interest']  += cash_int
            inv_totals['pik_interest']   += pik_int
            inv_totals['total_interest'] += tot_int
            inv_totals['oid']            += investor_oid
            inv_totals['fees']           += investor_fees
            inv_totals['total_income']   += total_income
            inv_totals['prepayments']    += prepay

            if first_beg_principal is None and beg_prin != 0:
                first_beg_principal = beg_prin
            if end_prin != 0:
                last_end_principal = end_prin

            # Alternate row shading
            row_fill = alt_fill if p_idx % 2 == 0 else None

            r = current_row
            ws.cell(r, COL["Period"],            period['period_number'])
            ws.cell(r, COL["Period Start"],      period['start_date'].strftime('%m/%d/%Y'))
            ws.cell(r, COL["Period End"],        period['end_date'].strftime('%m/%d/%Y'))
            ws.cell(r, COL["Ownership %"],       ownership / 100)
            ws.cell(r, COL["Beg. Principal"],    beg_prin)
            ws.cell(r, COL["Cash Interest"],     cash_int)
            ws.cell(r, COL["PIK Interest"],      pik_int)
            ws.cell(r, COL["Total Interest"],    tot_int)
            ws.cell(r, COL["OID Accreted"],      investor_oid)
            ws.cell(r, COL["Fee Income"],        investor_fees)
            ws.cell(r, COL["Total Income"],      total_income)
            ws.cell(r, COL["Principal Prepay"],  prepay)
            ws.cell(r, COL["End. Principal"],    end_prin)
            ws.cell(r, COL["Cumul. Interest"],   cumul_interest)
            ws.cell(r, COL["Cumul. OID"],        cumul_oid)
            ws.cell(r, COL["Cumul. Fees"],       cumul_fees)
            ws.cell(r, COL["Cumul. Total Inc."], cumul_total)

            # Formatting
            _fmt_pct(r, COL["Ownership %"])
            for col_label in ["Beg. Principal", "Cash Interest", "PIK Interest",
                               "Total Interest", "OID Accreted", "Fee Income",
                               "Total Income", "Principal Prepay", "End. Principal",
                               "Cumul. Interest", "Cumul. OID", "Cumul. Fees",
                               "Cumul. Total Inc."]:
                _fmt_currency(r, COL[col_label])

            if row_fill:
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(r, col_idx).fill = row_fill

            current_row += 1

        # ── Investor subtotal row ─────────────────────────────────────────────
        r = current_row
        ws.cell(r, COL["Period"],            "TOTAL")
        ws.cell(r, COL["Beg. Principal"],    first_beg_principal or 0.0)
        ws.cell(r, COL["Cash Interest"],     inv_totals['cash_interest'])
        ws.cell(r, COL["PIK Interest"],      inv_totals['pik_interest'])
        ws.cell(r, COL["Total Interest"],    inv_totals['total_interest'])
        ws.cell(r, COL["OID Accreted"],      inv_totals['oid'])
        ws.cell(r, COL["Fee Income"],        inv_totals['fees'])
        ws.cell(r, COL["Total Income"],      inv_totals['total_income'])
        ws.cell(r, COL["Principal Prepay"],  inv_totals['prepayments'])
        ws.cell(r, COL["End. Principal"],    last_end_principal or 0.0)

        for col_idx in range(1, len(columns) + 1):
            ws.cell(r, col_idx).font = sub_font
            ws.cell(r, col_idx).fill = sub_fill

        for col_label in ["Beg. Principal", "Cash Interest", "PIK Interest",
                           "Total Interest", "OID Accreted", "Fee Income",
                           "Total Income", "Principal Prepay", "End. Principal"]:
            _fmt_currency(r, COL[col_label])

        # Accumulate grand totals
        for k in grand:
            grand[k] += inv_totals.get(k, 0.0)

        current_row += 3   # blank spacer between investors

    # ── Grand Total block ─────────────────────────────────────────────────────
    r = current_row
    ws.cell(r, 1, "GRAND TOTAL — ALL INVESTORS").font = Font(bold=True, size=11, color="FFFFFF")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns))
    for col_idx in range(1, len(columns) + 1):
        ws.cell(r, col_idx).fill = grand_fill
    ws.row_dimensions[r].height = 18
    current_row += 1

    r = current_row
    labels = {
        COL["Period"]:           "TOTAL",
        COL["Cash Interest"]:    grand['cash_interest'],
        COL["PIK Interest"]:     grand['pik_interest'],
        COL["Total Interest"]:   grand['total_interest'],
        COL["OID Accreted"]:     grand['oid'],
        COL["Fee Income"]:       grand['fees'],
        COL["Total Income"]:     grand['total_income'],
        COL["Principal Prepay"]: grand['prepayments'],
    }
    for col_idx, value in labels.items():
        c = ws.cell(r, col_idx, value)
        c.font = grand_font
        c.fill = grand_fill
        if isinstance(value, float):
            c.number_format = CURRENCY

    # ── Column widths & freeze panes ─────────────────────────────────────────
    _set_col_widths()
    ws.freeze_panes = "B3"


def _create_payment_ledger_tab(wb, loan_id):
    """Create Payment Ledger tab."""
    from payments import load_payments
    
    ws = wb.create_sheet("Payment Ledger")
    
    # Headers
    headers = ["Date", "Type", "Amount", "Period", "Notes"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Load payments
    payments = load_payments(loan_id)
    
    # Data rows
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment['payment_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=payment['payment_type'])
        ws.cell(row=row, column=3, value=payment['amount'])
        ws.cell(row=row, column=4, value=payment.get('period', ''))
        ws.cell(row=row, column=5, value=payment.get('notes', ''))
        
        # Format currency
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_ownership_history_tab(wb, loan_id):
    """Create Ownership History tab."""
    from investors import load_investors
    
    ws = wb.create_sheet("Ownership History")
    
    # Headers
    headers = ["Effective Date", "Investor ID", "Investor Name", "Ownership %"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Load investors
    investors = load_investors(loan_id)
    
    # Data rows
    for row, investor in enumerate(investors, 2):
        ws.cell(row=row, column=1, value=investor['effective_date'].strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=investor['investor_id'])
        ws.cell(row=row, column=3, value=investor['investor_name'])
        ws.cell(row=row, column=4, value=investor['ownership_pct'] / 100)
        
        # Format percentage
        ws.cell(row=row, column=4).number_format = '0.00%'
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _create_reconciliation_tab(wb, loan_id, schedule):
    """Create Reconciliation tab."""
    from investors import load_investors
    from investor_allocation import allocate_period_to_investors
    
    ws = wb.create_sheet("Reconciliation")
    
    # Title
    ws['A1'] = "RECONCILIATION CHECKS"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    
    # Check 1: Interest allocations sum to period totals (penny-exact after rounding)
    ws[f'A{row}'] = "Check 1: Interest Allocations = Period Totals (Penny-Exact)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Period"
    ws[f'B{row}'] = "Interest Owed"
    ws[f'C{row}'] = "Total Allocated"
    ws[f'D{row}'] = "Remainder ($)"
    ws[f'E{row}'] = "Match"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(bold=True)
    row += 1

    for period in schedule:
        try:
            allocation = allocate_period_to_investors(loan_id, period)
            total_allocated = sum(inv['interest'] for inv in allocation['investor_allocations'])
            period_interest = period['interest_owed']
            remainder = round(total_allocated - period_interest, 2)
            # Penny-exact: tolerance is zero after Largest Remainder rounding
            match = abs(remainder) < 0.005

            ws[f'A{row}'] = f"Period {period['period_number']}"
            ws[f'B{row}'] = period_interest
            ws[f'C{row}'] = total_allocated
            ws[f'D{row}'] = remainder
            ws[f'E{row}'] = "✓" if match else "✗"
            ws[f'E{row}'].font = Font(color="00FF00" if match else "FF0000", bold=True)

            for col in ['B', 'C', 'D']:
                ws[f'{col}{row}'].number_format = '$#,##0.00'

            row += 1
        except:
            pass
    
    # Check 2: Fee allocations sum to period totals
    from fees import get_fees_for_period, get_fee_display_name
    from fee_allocation import allocate_fee_to_investors
    
    row += 2
    ws[f'A{row}'] = "Check 2: Fee Allocations = Fee Totals (Penny-Exact)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    ws[f'A{row}'] = "Period/Fee"
    ws[f'B{row}'] = "Total Fee"
    ws[f'C{row}'] = "Allocated"
    ws[f'D{row}'] = "Remainder ($)"
    ws[f'E{row}'] = "Match"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].font = Font(bold=True)
    row += 1

    fees_found = False
    for period in schedule:
        try:
            period_fees = get_fees_for_period(loan_id, period['period_number'])

            for fee in period_fees:
                fees_found = True
                fee_allocation = allocate_fee_to_investors(
                    loan_id=loan_id,
                    fee_date=fee['fee_date'],
                    fee_amount=fee['amount'],
                    fee_type=fee['fee_type']
                )

                total_allocated = sum(inv['fee_share'] for inv in fee_allocation['investor_allocations'])
                remainder = round(total_allocated - fee['amount'], 2)
                match = abs(remainder) < 0.005  # penny-exact after Largest Remainder rounding

                ws[f'A{row}'] = f"P{period['period_number']} - {get_fee_display_name(fee['fee_type'])}"
                ws[f'B{row}'] = fee['amount']
                ws[f'C{row}'] = total_allocated
                ws[f'D{row}'] = remainder
                ws[f'E{row}'] = "✓" if match else "✗"
                ws[f'E{row}'].font = Font(color="00FF00" if match else "FF0000", bold=True)

                for col in ['B', 'C', 'D']:
                    ws[f'{col}{row}'].number_format = '$#,##0.00'

                row += 1
        except:
            pass

    if not fees_found:
        ws[f'A{row}'] = "No fees to reconcile"

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 10

def _create_fee_income_tab(wb, loan_id):
    """Create Fee Income tab."""
    from fees import load_fees, get_fee_display_name
    
    ws = wb.create_sheet("Fee Income")
    
    # Headers
    headers = [
        "Fee Date", "Period", "Fee Type", "Amount", "Cash/PIK", "Description"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Load fees
    fees = load_fees(loan_id)
    
    if not fees:
        # No fees - add a note
        ws.cell(row=2, column=1, value="No fees recorded for this loan")
        ws.merge_cells('A2:F2')
    else:
        # Data rows
        for row, fee in enumerate(fees, 2):
            ws.cell(row=row, column=1, value=fee['fee_date'].strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=fee['period_number'] if fee['period_number'] else 'N/A')
            ws.cell(row=row, column=3, value=get_fee_display_name(fee['fee_type']))
            ws.cell(row=row, column=4, value=fee['amount'])
            ws.cell(row=row, column=5, value=fee['cash_or_pik'].upper())
            ws.cell(row=row, column=6, value=fee['description'])
            
            # Format currency
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
        
        # Add total row
        total_row = len(fees) + 2
        ws.cell(row=total_row, column=3, value="TOTAL:")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=sum(f['amount'] for f in fees))
        ws.cell(row=total_row, column=4).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=4).font = Font(bold=True)
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Description column wider
    ws.column_dimensions['F'].width = 40